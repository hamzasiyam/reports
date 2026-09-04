import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
import json

PROFILE_DIR = "./profiles"
PROFILE_FILE = os.path.join(PROFILE_DIR, "profiles.json")

def parse_csv_to_dicts(file_path):
    sections = []
    current_section = None
    with open(file_path, 'r') as file:
        for line in file:
            line = line.strip()
            if line.startswith('"Metric"'):
                if current_section:
                    sections.append(current_section)
                headers = line.split(',')[1:]  # Skip the "Metric" column
                headers = [header.strip('"') for header in headers]
                current_section = {"headers": headers, "data": []}
            elif current_section:
                data = line.split(',')[1:]  # Skip the "Metric" column
                data = [d.strip('"') for d in data if d.strip('"')]
                if data:
                    current_section["data"].append(data)
        if current_section:
            sections.append(current_section)
    return sections

def auto_size_columns(ws):
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

def create_excel_report(sections, output_path, preparer_name, preparer_email, preparer_phone, stats_month_year, prep_date):
    wb = Workbook()
    
    # Add a sheet for the spreadsheet information
    info_sheet = wb.active
    info_sheet.title = "Report Information"
    info_sheet.append(["Report Information", ""])
    info_sheet.append(["Preparer", preparer_name])
    info_sheet.append(["Preparer's Email", preparer_email])
    info_sheet.append(["Preparer's Office Phone", preparer_phone])
    info_sheet.append(["Stats Month and Year", stats_month_year])
    info_sheet.append(["Preparation Date", prep_date])
    info_sheet.append(["Note", f"The stats are for the entire month of {stats_month_year}."])
    auto_size_columns(info_sheet)

    # Calculate sum of Sessions with new users and Sessions with returning users
    total_sessions = 0
    for section in sections:
        if "Users overview" in section["headers"]:
            users_overview_data = section["data"]
            for row in users_overview_data:
                if len(row) > 1:
                    if "Sessions with new users" in row[0]:
                        total_sessions += int(row[1])
                    elif "Sessions with returning users" in row[0]:
                        total_sessions += int(row[1])
            break

    # Add a new sheet with the total sessions
    summary_sheet = wb.create_sheet(title="Sessions Summary", index=1)
    summary_sheet.append(["Sessions Summary", ""])
    summary_sheet.append(["Total User Sessions for the Month", total_sessions])
    auto_size_columns(summary_sheet)

    # Create sheets for the metric data
    for section in sections:
        headers = section['headers']
        sheet_name = headers[0] if len(headers) > 0 else 'Metric'
        ws = wb.create_sheet(title=sheet_name)
        ws.append(headers)
        for row in section['data']:
            ws.append(row)
        auto_size_columns(ws)
    
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # Ensure the output directory exists
    output_dir = "reports/processed"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    output_path = os.path.join(output_dir, output_path)
    wb.save(output_path)

def save_profile(name, email, phone):
    if not os.path.exists(PROFILE_DIR):
        os.makedirs(PROFILE_DIR)

    profiles = load_profiles()
    profiles[name] = {
        "email": email,
        "phone": phone
    }
    
    with open(PROFILE_FILE, "w") as f:
        json.dump(profiles, f, indent=4)

def load_profiles():
    if os.path.exists(PROFILE_FILE):
        with open(PROFILE_FILE, "r") as f:
            return json.load(f)
    return {}

def select_profile(profile_name_var, name_var, email_var, phone_var):
    profiles = load_profiles()
    profile_name = profile_name_var.get()
    if profile_name in profiles:
        profile = profiles[profile_name]
        name_var.set(profile_name)
        email_var.set(profile["email"])
        phone_var.set(profile["phone"])

def edit_profile():
    selected_profile = profile_name_var.get()
    if not selected_profile:
        messagebox.showerror("Error", "Please select a profile to edit")
        return

    profiles = load_profiles()
    profile = profiles.get(selected_profile, {})

    profile_window = tk.Toplevel(root)
    profile_window.title("Edit Profile")
    profile_window.geometry("400x200")
    
    tk.Label(profile_window, text="Preparer Name").grid(row=0, column=0, pady=10, sticky=tk.W)
    profile_name = tk.Entry(profile_window, width=50)
    profile_name.insert(0, selected_profile)
    profile_name.grid(row=0, column=1, pady=10)
    
    tk.Label(profile_window, text="Preparer Email").grid(row=1, column=0, pady=10, sticky=tk.W)
    profile_email = tk.Entry(profile_window, width=50)
    profile_email.insert(0, profile.get("email", ""))
    profile_email.grid(row=1, column=1, pady=10)
    
    tk.Label(profile_window, text="Preparer Phone").grid(row=2, column=0, pady=10, sticky=tk.W)
    profile_phone = tk.Entry(profile_window, width=50)
    profile_phone.insert(0, profile.get("phone", ""))
    profile_phone.grid(row=2, column=1, pady=10)
    
    def save_edited_profile():
        name = profile_name.get()
        email = profile_email.get()
        phone = profile_phone.get()
        profiles[name] = {
            "email": email,
            "phone": phone
        }
        if selected_profile != name:
            del profiles[selected_profile]
        with open(PROFILE_FILE, "w") as f:
            json.dump(profiles, f, indent=4)
        profile_window.destroy()
        load_profile_names()

    tk.Button(profile_window, text="Save Profile", command=save_edited_profile).grid(row=3, columnspan=2, pady=10)

def delete_profile():
    selected_profile = profile_name_var.get()
    if not selected_profile:
        messagebox.showerror("Error", "Please select a profile to delete")
        return

    profiles = load_profiles()
    if selected_profile in profiles:
        del profiles[selected_profile]
        with open(PROFILE_FILE, "w") as f:
            json.dump(profiles, f, indent=4)
        load_profile_names()
        messagebox.showinfo("Success", f"Profile '{selected_profile}' deleted successfully")

def load_profile_names():
    profiles = load_profiles()
    profile_names = list(profiles.keys())
    profile_name_combobox['values'] = profile_names

def main():
    global root, profile_name_var, profile_name_combobox  # Declare global variables
    def select_file():
        csv_path.set(filedialog.askopenfilename(title="Select the CSV file", filetypes=[("CSV files", "*.csv")]))
    
    def create_profile():
        profile_window = tk.Toplevel(root)
        profile_window.title("Create Profile")
        profile_window.geometry("400x200")
        
        tk.Label(profile_window, text="Preparer Name").grid(row=0, column=0, pady=10, sticky=tk.W)
        profile_name = tk.Entry(profile_window, width=50)
        profile_name.grid(row=0, column=1, pady=10)
        
        tk.Label(profile_window, text="Preparer Email").grid(row=1, column=0, pady=10, sticky=tk.W)
        profile_email = tk.Entry(profile_window, width=50)
        profile_email.grid(row=1, column=1, pady=10)
        
        tk.Label(profile_window, text="Preparer Phone").grid(row=2, column=0, pady=10, sticky=tk.W)
        profile_phone = tk.Entry(profile_window, width=50)
        profile_phone.grid(row=2, column=1, pady=10)
        
        def save_new_profile():
            name = profile_name.get()
            email = profile_email.get()
            phone = profile_phone.get()
            save_profile(name, email, phone)
            profile_window.destroy()
            load_profile_names()
        
        tk.Button(profile_window, text="Save Profile", command=save_new_profile).grid(row=3, columnspan=2, pady=10)
    
    def submit():
        path = csv_path.get()
        if not path:
            messagebox.showerror("Error", "Please select a CSV file")
            return

        preparer_name = entry_name.get()
        preparer_email = entry_email.get()
        preparer_phone = entry_phone.get()
        stats_month_year = entry_month_year.get()
        prep_date = entry_prep_date.get()

        if not all([preparer_name, preparer_email, preparer_phone, stats_month_year, prep_date]):
            messagebox.showerror("Error", "Please fill in all fields")
            return

        sections = parse_csv_to_dicts(path)
        excel_path = f"Website Performance Insights {stats_month_year} Spreadsheet by {preparer_name}.xlsx"
        create_excel_report(sections, excel_path, preparer_name, preparer_email, preparer_phone, stats_month_year, prep_date)
        messagebox.showinfo("Success", f"Excel report created: {excel_path}")

    def select_profile_callback(event):
        select_profile(profile_name_var, entry_name_var, entry_email_var, entry_phone_var)

    root = tk.Tk()
    root.title("Excel Report Generator")
    root.geometry("950x400")

    style = ttk.Style()
    style.configure("TLabel", font=("Helvetica", 12))
    style.configure("TButton", font=("Helvetica", 12))
    style.configure("TEntry", font=("Helvetica", 12))

    main_frame = ttk.Frame(root, padding="10")
    main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

    ttk.Label(main_frame, text="Select CSV File").grid(row=0, column=0, pady=10, sticky=tk.W)
    csv_path = tk.StringVar()
    ttk.Entry(main_frame, textvariable=csv_path, width=50).grid(row=0, column=1, pady=10, padx=10)
    ttk.Button(main_frame, text="Browse", command=select_file).grid(row=0, column=2, pady=10)

    ttk.Label(main_frame, text="Profile Name").grid(row=1, column=0, pady=10, sticky=tk.W)
    profile_name_var = tk.StringVar()
    profile_name_combobox = ttk.Combobox(main_frame, textvariable=profile_name_var, state='readonly')
    profile_name_combobox.grid(row=1, column=1, pady=10, padx=10)
    profile_name_combobox.bind("<<ComboboxSelected>>", select_profile_callback)
    
    ttk.Button(main_frame, text="Create Profile", command=create_profile).grid(row=1, column=2, pady=10)
    ttk.Button(main_frame, text="Edit Profile", command=edit_profile).grid(row=1, column=3, pady=10)
    ttk.Button(main_frame, text="Delete Profile", command=delete_profile).grid(row=1, column=4, pady=10)

    ttk.Label(main_frame, text="Preparer Name").grid(row=2, column=0, pady=10, sticky=tk.W)
    entry_name_var = tk.StringVar()
    entry_name = ttk.Entry(main_frame, textvariable=entry_name_var, width=50)
    entry_name.grid(row=2, column=1, pady=10, padx=10, columnspan=2)

    ttk.Label(main_frame, text="Preparer Email").grid(row=3, column=0, pady=10, sticky=tk.W)
    entry_email_var = tk.StringVar()
    entry_email = ttk.Entry(main_frame, textvariable=entry_email_var, width=50)
    entry_email.grid(row=3, column=1, pady=10, padx=10, columnspan=2)

    ttk.Label(main_frame, text="Preparer Phone").grid(row=4, column=0, pady=10, sticky=tk.W)
    entry_phone_var = tk.StringVar()
    entry_phone = ttk.Entry(main_frame, textvariable=entry_phone_var, width=50)
    entry_phone.grid(row=4, column=1, pady=10, padx=10, columnspan=2)

    ttk.Label(main_frame, text="Stats Month and Year (e.g., May 2024)").grid(row=5, column=0, pady=10, sticky=tk.W)
    entry_month_year = ttk.Entry(main_frame, width=50)
    entry_month_year.grid(row=5, column=1, pady=10, padx=10, columnspan=2)

    ttk.Label(main_frame, text="Preparation Date (e.g., June 3 2024)").grid(row=6, column=0, pady=10, sticky=tk.W)
    entry_prep_date = ttk.Entry(main_frame, width=50)
    entry_prep_date.grid(row=6, column=1, pady=10, padx=10, columnspan=2)

    ttk.Button(main_frame, text="Generate Excel Sheet", command=submit).grid(row=7, column=0, columnspan=3, pady=20)

    load_profile_names()

    root.mainloop()

if __name__ == "__main__":
    main()
